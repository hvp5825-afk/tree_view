import csv
import io
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.http import HttpResponse
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from helpdesk.models import SendHelpRequest
from pmf.models import PMFRequest

User = get_user_model()


@login_required
def reports_index(request):
    if not request.user.is_staff:
        return redirect('dashboard:user_dashboard')
    return render(request, 'reports/index.html')


@login_required
def export_members_excel(request):
    if not request.user.is_staff:
        return HttpResponse(status=403)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Members'
    ws.append(['Sl', 'Member ID', 'Username', 'Email', 'Mobile', 'Sponsor ID', 'Status', 'Joined'])
    for i, u in enumerate(User.objects.filter(is_staff=False).order_by('created_at'), 1):
        ws.append([i, u.member_id, u.username, u.email, u.mobile, u.sponsor_id or '', u.status, str(u.created_at.date())])
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=members.xlsx'
    wb.save(resp)
    return resp


@login_required
def export_members_csv(request):
    if not request.user.is_staff:
        return HttpResponse(status=403)
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename=members.csv'
    writer = csv.writer(resp)
    writer.writerow(['Sl', 'Member ID', 'Username', 'Email', 'Mobile', 'Sponsor ID', 'Status', 'Joined'])
    for i, u in enumerate(User.objects.filter(is_staff=False).order_by('created_at'), 1):
        writer.writerow([i, u.member_id, u.username, u.email, u.mobile, u.sponsor_id or '', u.status, str(u.created_at.date())])
    return resp


@login_required
def export_members_pdf(request):
    if not request.user.is_staff:
        return HttpResponse(status=403)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setFont('Helvetica-Bold', 14)
    p.drawString(200, 800, 'Members Report')
    p.setFont('Helvetica', 9)
    y = 770
    headers = ['#', 'Member ID', 'Username', 'Email', 'Mobile', 'Status']
    x_positions = [30, 50, 120, 200, 320, 430]
    for x, h in zip(x_positions, headers):
        p.drawString(x, y, h)
    y -= 15
    for i, u in enumerate(User.objects.filter(is_staff=False).order_by('created_at'), 1):
        if y < 50:
            p.showPage()
            y = 800
        row = [str(i), u.member_id, u.username, u.email, u.mobile, u.status]
        for x, val in zip(x_positions, row):
            p.drawString(x, y, val[:20])
        y -= 14
    p.save()
    buffer.seek(0)
    resp = HttpResponse(buffer, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename=members.pdf'
    return resp


@login_required
def export_help_excel(request):
    if not request.user.is_staff:
        return HttpResponse(status=403)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Help Requests'
    ws.append(['Sl', 'Request By', 'Request To', 'Amount', 'Method', 'Trans No', 'Status', 'Date'])
    for i, s in enumerate(SendHelpRequest.objects.select_related('user', 'request_to').order_by('-created_at'), 1):
        ws.append([i, s.user.member_id, s.request_to.member_id if s.request_to else '', str(s.amount), s.payment_method, s.transaction_no, s.status, str(s.created_at.date())])
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=help_requests.xlsx'
    wb.save(resp)
    return resp


@login_required
def export_pmf_excel(request):
    if not request.user.is_staff:
        return HttpResponse(status=403)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PMF Requests'
    ws.append(['Sl', 'Member ID', 'Name', 'Amount', 'Trans No', 'Level', 'Status', 'Date'])
    for i, pmf in enumerate(PMFRequest.objects.select_related('user').order_by('-created_at'), 1):
        ws.append([i, pmf.user.member_id, pmf.user.get_full_name(), str(pmf.amount), pmf.transaction_no, pmf.level, pmf.status, str(pmf.created_at.date())])
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=pmf_requests.xlsx'
    wb.save(resp)
    return resp
